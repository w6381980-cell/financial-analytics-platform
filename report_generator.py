import pandas as pd
from sqlalchemy import create_engine
import urllib
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Image, Spacer
from reportlab.lib.styles import getSampleStyleSheet
import openpyxl
import openpyxl.styles as xl_styles          # ← fix: alag naam
from openpyxl.chart import BarChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

# ── CONNECTION ────────────────────────────────────────────────────
params = urllib.parse.quote_plus(
    "DRIVER={ODBC Driver 17 for SQL Server};"
    "SERVER=LAPTOP-LG4BEQ1J\\SQLEXPRESS;"
    "DATABASE=FinanceDB;"
    "Trusted_Connection=yes;"
)
engine = create_engine(f"mssql+pyodbc:///?odbc_connect={params}")
print("✓ SQL Server connected")

# ── DATA LOAD ─────────────────────────────────────────────────────
df_pl = pd.read_sql(
    "SELECT * FROM vw_pl_summary ORDER BY ticker, quarter",
    engine
)
df_proj = pd.read_sql("""
    SELECT ticker, quarter, projected_rev, lower_bound, upper_bound, model_used
    FROM projections
    WHERE model_used = 'LinearRegression'
    ORDER BY ticker, quarter
""", engine)
df_stocks = pd.read_sql("""
    SELECT [date], ticker, [close]
    FROM stocks
    WHERE [date] >= '2023-01-01'
    ORDER BY [date]
""", engine)

print(f"✓ P&L rows    : {len(df_pl)}")
print(f"✓ Projections : {len(df_proj)}")
print(f"✓ Stocks      : {len(df_stocks)}")

# ── KPI SUMMARY ───────────────────────────────────────────────────
kpi = df_pl[df_pl['quarter'].str.startswith('2023')]
kpi_summary = kpi.groupby('ticker').agg(
    Avg_Revenue    =('revenue',         'mean'),
    Avg_Net_Profit =('net_profit',      'mean'),
    Net_Margin     =('net_margin_pct',  'mean'),
    Gross_Margin   =('gross_margin_pct','mean')
).round(2).reset_index()

# ════════════════════════════════════════════════════════════════════
# PART A: PDF REPORT
# ════════════════════════════════════════════════════════════════════
print("\n── Generating PDF ──")

doc      = SimpleDocTemplate(
    "report.pdf", pagesize=A4,
    topMargin=2*cm, bottomMargin=2*cm,
    leftMargin=1.5*cm, rightMargin=1.5*cm
)
rl_styles = getSampleStyleSheet()            # ← reportlab styles
story     = []

# ── Title ─────────────────────────────────────────────────────────
story.append(Spacer(1, 1*cm))
story.append(Paragraph("<b>Financial Analytics Report</b>", rl_styles['Title']))
story.append(Paragraph(
    "Finance &amp; Stock Data — Intermediate Analytics Project",
    rl_styles['Heading3']
))
story.append(Paragraph(
    f"Generated: {datetime.now().strftime('%d %B %Y, %H:%M')}",
    rl_styles['Normal']
))
story.append(Paragraph(
    "Companies: ALPHACORP · BETAFIN · DELTATECH · EPSILONBK · GAMMAINV",
    rl_styles['Normal']
))
story.append(Spacer(1, 0.8*cm))

# ── Section 1: KPI Summary Table ──────────────────────────────────
story.append(Paragraph(
    "<b>1. P&amp;L Summary — FY 2023 (Average per Quarter)</b>",
    rl_styles['Heading2']
))
story.append(Spacer(1, 0.3*cm))

table_data = [['Company', 'Avg Revenue', 'Avg Net Profit', 'Net Margin %', 'Gross Margin %']]
for _, row in kpi_summary.iterrows():
    table_data.append([
        row['ticker'],
        f"Rs.{row['Avg_Revenue']/1e6:.1f}M",
        f"Rs.{row['Avg_Net_Profit']/1e6:.1f}M",
        f"{row['Net_Margin']:.1f}%",
        f"{row['Gross_Margin']:.1f}%"
    ])

tbl = Table(table_data, colWidths=[3.5*cm, 3.5*cm, 3.5*cm, 3*cm, 3.5*cm])
tbl.setStyle(TableStyle([
    ('BACKGROUND',     (0, 0), (-1, 0),  colors.HexColor('#1565C0')),
    ('TEXTCOLOR',      (0, 0), (-1, 0),  colors.white),
    ('FONTNAME',       (0, 0), (-1, 0),  'Helvetica-Bold'),
    ('FONTSIZE',       (0, 0), (-1, 0),  10),
    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#E3F2FD')]),
    ('FONTSIZE',       (0, 1), (-1, -1), 9),
    ('GRID',           (0, 0), (-1, -1), 0.5, colors.HexColor('#BBDEFB')),
    ('ALIGN',          (1, 1), (-1, -1), 'RIGHT'),
    ('ALIGN',          (0, 0), (0,  -1), 'LEFT'),
    ('TOPPADDING',     (0, 0), (-1, -1), 6),
    ('BOTTOMPADDING',  (0, 0), (-1, -1), 6),
]))
story.append(tbl)
story.append(Spacer(1, 0.6*cm))

# ── Section 2: Projections Table ──────────────────────────────────
story.append(Paragraph(
    "<b>2. Revenue Projections — FY 2024 (Linear Regression)</b>",
    rl_styles['Heading2']
))
story.append(Spacer(1, 0.3*cm))

proj_table_data = [['Company', 'Quarter', 'Projected Rev', 'Lower Bound', 'Upper Bound']]
for _, row in df_proj.iterrows():
    proj_table_data.append([
        row['ticker'],
        row['quarter'],
        f"Rs.{row['projected_rev']/1e6:.1f}M",
        f"Rs.{row['lower_bound']/1e6:.1f}M",
        f"Rs.{row['upper_bound']/1e6:.1f}M",
    ])

tbl2 = Table(proj_table_data, colWidths=[3.5*cm, 2.5*cm, 3.5*cm, 3.5*cm, 3.5*cm])
tbl2.setStyle(TableStyle([
    ('BACKGROUND',     (0, 0), (-1, 0),  colors.HexColor('#E65100')),
    ('TEXTCOLOR',      (0, 0), (-1, 0),  colors.white),
    ('FONTNAME',       (0, 0), (-1, 0),  'Helvetica-Bold'),
    ('FONTSIZE',       (0, 0), (-1, 0),  10),
    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#FBE9E7')]),
    ('FONTSIZE',       (0, 1), (-1, -1), 9),
    ('GRID',           (0, 0), (-1, -1), 0.5, colors.HexColor('#FFCCBC')),
    ('ALIGN',          (2, 1), (-1, -1), 'RIGHT'),
    ('TOPPADDING',     (0, 0), (-1, -1), 5),
    ('BOTTOMPADDING',  (0, 0), (-1, -1), 5),
]))
story.append(tbl2)
story.append(Spacer(1, 0.6*cm))

# ── Section 3: Charts ─────────────────────────────────────────────
chart_items = [
    ('charts/revenue_trend.png',     '3. Quarterly Revenue Trend'),
    ('charts/pl_waterfall.png',      '4. P&L Comparison — Q4 2023'),
    ('charts/net_margin.png',        '5. Net Profit Margin by Company'),
    ('charts/stock_performance.png', '6. Normalized Stock Performance 2023'),
    ('charts/forecast.png',          '7. Revenue Forecast — 2024'),
]
for chart_path, caption in chart_items:
    story.append(Paragraph(f"<b>{caption}</b>", rl_styles['Heading2']))
    story.append(Image(chart_path, width=16*cm, height=8*cm))
    story.append(Spacer(1, 0.5*cm))

doc.build(story)
print("✓ report.pdf generated")

# ════════════════════════════════════════════════════════════════════
# PART B: EXCEL REPORT
# ════════════════════════════════════════════════════════════════════
print("\n── Generating Excel ──")

wb = openpyxl.Workbook()

# Style constants — xl_styles use kar rahe hain (clash nahi hoga)
HEADER_FILL = xl_styles.PatternFill("solid", fgColor="1565C0")
HEADER_FONT = xl_styles.Font(color="FFFFFF", bold=True)
ORANGE_FILL = xl_styles.PatternFill("solid", fgColor="E65100")
ALT_FILL    = xl_styles.PatternFill("solid", fgColor="E3F2FD")

def style_header(ws, fill, font):
    for cell in ws[1]:
        cell.fill      = fill
        cell.font      = font
        cell.alignment = xl_styles.Alignment(horizontal='center')

# ── Sheet 1: P&L Data ─────────────────────────────────────────────
ws1 = wb.active
ws1.title = "P&L Data"
for r in dataframe_to_rows(df_pl, index=False, header=True):
    ws1.append(r)
style_header(ws1, HEADER_FILL, HEADER_FONT)
for i, row in enumerate(ws1.iter_rows(min_row=2), start=2):
    if i % 2 == 0:
        for cell in row:
            cell.fill = ALT_FILL
ws1.column_dimensions['A'].width = 12
ws1.column_dimensions['B'].width = 12

# ── Sheet 2: KPI Summary ──────────────────────────────────────────
ws2 = wb.create_sheet("KPI Summary")
for r in dataframe_to_rows(kpi_summary, index=False, header=True):
    ws2.append(r)
style_header(ws2, HEADER_FILL, HEADER_FONT)
for col in ws2.columns:
    ws2.column_dimensions[col[0].column_letter].width = 18

# ── Sheet 3: Projections ──────────────────────────────────────────
ws3 = wb.create_sheet("Projections 2024")
for r in dataframe_to_rows(df_proj, index=False, header=True):
    ws3.append(r)
style_header(ws3, ORANGE_FILL, HEADER_FONT)
for col in ws3.columns:
    ws3.column_dimensions[col[0].column_letter].width = 18

# ── Sheet 4: Revenue Bar Chart ────────────────────────────────────
ws4 = wb.create_sheet("Revenue Chart")
ws4.append(['Company', 'Avg Revenue (M)'])
for _, row in kpi_summary.iterrows():
    ws4.append([row['ticker'], round(row['Avg_Revenue'] / 1e6, 2)])
style_header(ws4, HEADER_FILL, HEADER_FONT)

chart = BarChart()
chart.title        = "Average Revenue by Company (Rs. Million)"
chart.y_axis.title = "Revenue (Rs.M)"
chart.x_axis.title = "Company"
chart.style        = 10
data_ref = Reference(ws4, min_col=2, min_row=1, max_row=6)
cats_ref = Reference(ws4, min_col=1, min_row=2, max_row=6)
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)
ws4.add_chart(chart, "D2")

# ── Sheet 5: Stock Data ───────────────────────────────────────────
ws5 = wb.create_sheet("Stock Data")
for r in dataframe_to_rows(df_stocks, index=False, header=True):
    ws5.append(r)
style_header(ws5, HEADER_FILL, HEADER_FONT)
ws5.column_dimensions['A'].width = 14
ws5.column_dimensions['B'].width = 14
ws5.column_dimensions['C'].width = 14

wb.save("report.xlsx")
print("✓ report.xlsx generated")

print("\n" + "="*50)
print("✓ ALL DONE!")
print("  report.pdf  → D:\\python\\finance_project\\report.pdf")
print("  report.xlsx → D:\\python\\finance_project\\report.xlsx")
print("="*50)